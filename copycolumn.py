# Load Pandas and openpyxl
import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel('Wave.xlsx')
# Reduce dataframe to target columns (by filtering on column names)
df = df[['fruit', 'days']]

#Copying the df output to excel
out = pd.ExcelWriter('Book1.xlsx')
df.to_excel(out, index=False)
out.save()
print('1')

#Loading the required Worksheets
src_wb = load_workbook('Book1.xlsx')
dest_wb = load_workbook('Blueprints_Input_Sample.xlsx')
src_sheet = src_wb.get_sheet_by_name('Sheet1')
dest_sheet = dest_wb.get_sheet_by_name('Blueprints_Input_Sample')

#Copying data from Src excel to required excel
for i in range(1, src_sheet.max_row+1):
    for j in range(1, src_sheet.max_column+1):
        dest_sheet.cell(row=i+1, column=j+2).value = src_sheet.cell(row=i+1, column=j).value

#src_wb.save('source.xlsx')
dest_wb.save('Blueprints.xlsx')








